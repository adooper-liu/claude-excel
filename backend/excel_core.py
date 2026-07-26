#!/usr/bin/env python3
"""
excel_core.py — Stable Python utility library for Claude Code + Excel workflows.

All functions are self-contained: they return DataFrames or dicts, never print.
Claude Code handles display and user interaction. Each function is also a CLI
subcommand for standalone use and debugging.

Usage:
    python tools/excel_core.py describe sales.xlsx
    python tools/excel_core.py clean dirty.xlsx --output cleaned.xlsx
    python tools/excel_core.py compare v1.xlsx v2.xlsx --key "订单号"
    python tools/excel_core.py profile sales.xlsx --columns "金额,数量"
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from typing import Any, Optional

import numpy as np
import pandas as pd

# ── Section 0: Constants ───────────────────────────────────────────────

LARGE_ROW_THRESHOLD = 100_000
LARGE_COL_THRESHOLD = 50
LARGE_FILE_MB = 50

# ── Section 1: Reading ──────────────────────────────────────────────────


def read_excel(
    path: str,
    sheet: Any = 0,
    header: bool | int = True,
    nrows: Optional[int] = None,
    usecols: Any = None,
    password: Optional[str] = None,
) -> pd.DataFrame | dict[str, pd.DataFrame]:
    """Read Excel (.xlsx/.xls) or CSV into a DataFrame.

    Args:
        path: Path to the file (absolute recommended).
        sheet: Sheet name (str), zero-based index (int), None for all sheets,
               or a list of names/indices.
        header: Row to use as column names (True=row 0, False=no header, int=N).
        nrows: Max rows to read (None=all). Use for sampling large files.
        usecols: Columns to read (list of names, indices, or Excel ranges).
        password: Password for encrypted .xlsx files.

    Returns:
        A single DataFrame (if one sheet) or dict[str, DataFrame] (if
        sheet=None or sheet is a list).
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        if isinstance(sheet, (int, str)) and sheet not in (0, "Sheet1"):
            # CSV has no sheets; pass through silently
            pass
        # Pass valid kwargs only; openpyxl params (password) are ignored for CSV
        kwargs: dict[str, Any] = {"nrows": nrows, "usecols": usecols}
        if isinstance(header, int):
            kwargs["header"] = header
        elif not header:
            kwargs["header"] = None
        result = pd.read_csv(path, **kwargs)
        if isinstance(sheet, (list, type(None))) and sheet != 0:
            result = {"Sheet1": result}
        return result

    # Normalize header: newer pandas rejects bool; True→0, False→None
    _header: Any = 0 if header is True else (None if header is False else header)
    read_kwargs: dict[str, Any] = {"header": _header, "nrows": nrows, "usecols": usecols}
    if password:
        read_kwargs["password"] = password
    else:
        read_kwargs["engine"] = "openpyxl"

    result = pd.read_excel(path, sheet_name=sheet, **read_kwargs)  # type: ignore[arg-type]
    return result


def describe(path: str) -> dict[str, Any]:
    """Structural summary of a workbook — no AI needed.

    Returns a dict ready for JSON serialization:
        {
          "file_name": str,
          "file_size_mb": float,
          "sheet_count": int,
          "sheets": {
            "<name>": {
              "rows": int, "cols": int,
              "column_names": [str, ...],
              "dtypes": {col: dtype_str, ...},
              "null_counts": {col: int, ...},
              "null_pct": {col: float, ...},
              "sample_rows": [{col: val}, ...]  # first 3
            }
          },
          "warnings": [str, ...]  # merged cells, large size, etc.
        }
    """
    file_size = os.path.getsize(path) / (1024 * 1024)
    warnings: list[str] = []

    if file_size > LARGE_FILE_MB:
        warnings.append(
            f"File is {file_size:.1f} MB (> {LARGE_FILE_MB} MB). "
            f"Consider sampling with read_excel(..., nrows=N)."
        )

    # Detect merged cells via openpyxl before pandas read
    merged_cells: dict[str, list[str]] = {}
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                if ws.merged_cells.ranges:
                    merged_cells[sn] = [str(r) for r in ws.merged_cells.ranges]
            wb.close()
        except Exception:
            pass  # openpyxl inspection is best-effort

    all_sheets = read_excel(path, sheet=None)
    if not isinstance(all_sheets, dict):
        all_sheets = {"Sheet1": all_sheets}

    sheets_info: dict[str, Any] = {}
    total_rows = 0
    for name, df in all_sheets.items():
        rows, cols = df.shape
        total_rows += rows
        dtypes_map = {str(c): str(df[c].dtype) for c in df.columns}
        null_counts = {str(c): int(df[c].isna().sum()) for c in df.columns}
        null_pct = {
            str(c): round(df[c].isna().sum() / max(rows, 1) * 100, 1)
            for c in df.columns
        }
        # Best-effort datetime detection for display
        col_names = [str(c) for c in df.columns]
        sample = df.head(3).where(df.notna(), None).to_dict(orient="records")

        sheets_info[name] = {
            "rows": rows,
            "cols": cols,
            "column_names": col_names,
            "dtypes": dtypes_map,
            "null_counts": null_counts,
            "null_pct": null_pct,
            "sample_rows": sample,
        }

    if total_rows > LARGE_ROW_THRESHOLD:
        warnings.append(
            f"Total rows ({total_rows}) > {LARGE_ROW_THRESHOLD}. "
            f"Recommend sampling for analysis."
        )
    if any(len(s["column_names"]) > LARGE_COL_THRESHOLD for s in sheets_info.values()):
        warnings.append(
            f"Some sheets have > {LARGE_COL_THRESHOLD} columns. "
            f"Context window may overflow — narrow your selection."
        )
    if merged_cells:
        flat = []
        for sn, ranges in merged_cells.items():
            flat.append(f"  {sn}: {', '.join(ranges[:10])}")
            if len(ranges) > 10:
                flat[-1] += f" ... +{len(ranges) - 10} more"
        warnings.append(f"Merged cells detected:\n" + "\n".join(flat))

    return {
        "file_name": os.path.basename(path),
        "file_size_mb": round(file_size, 2),
        "sheet_count": len(sheets_info),
        "sheets": sheets_info,
        "warnings": warnings,
    }


# ── Section 2: Cleaning ──────────────────────────────────────────────────


def trim_all_text(df: pd.DataFrame) -> pd.DataFrame:
    """Strip leading/trailing whitespace from all string/object columns."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    return df


def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names: strip, replace spaces with _, remove special chars.
    Preserves non-ASCII characters (Chinese, Japanese, etc.)."""
    df = df.copy()
    mapping = {}
    for c in df.columns:
        s = str(c).strip()
        # Replace whitespace sequences with single underscore
        s = "_".join(s.split())
        # Remove characters that cause trouble in code (but keep Unicode)
        s = s.replace('"', "").replace("'", "").replace("`", "")
        mapping[c] = s
    return df.rename(columns=mapping)


def normalize_dates(
    df: pd.DataFrame,
    columns: list[str],
    fmt: Optional[str] = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Parse date columns with format auto-detection.

    Returns (cleaned_df, report) where report has:
        {col: {"parsed": N, "failed": N, "failed_values": [...]}}
    """
    df = df.copy()
    report: dict[str, Any] = {}
    for col in columns:
        if col not in df.columns:
            report[col] = {"error": "column not found"}
            continue
        original = df[col].copy()
        try:
            parsed = pd.to_datetime(original, format=fmt, errors="coerce")
        except Exception:
            parsed = pd.to_datetime(original, errors="coerce")
        failed_mask = parsed.isna() & original.notna()
        failed_vals = (
            original[failed_mask].unique().tolist()[:20] if failed_mask.any() else []
        )
        report[col] = {
            "parsed": int((~parsed.isna()).sum()),
            "failed": int(failed_mask.sum()),
            "failed_values": failed_vals,
        }
        df[col] = parsed
    return df, report


def detect_duplicates(
    df: pd.DataFrame,
    subset: Optional[list[str]] = None,
) -> pd.DataFrame:
    """Return duplicated rows with a 'duplicate_group' column for tracing."""
    dup_mask = df.duplicated(subset=subset, keep=False)
    if not dup_mask.any():
        return pd.DataFrame(columns=list(df.columns) + ["duplicate_group"])
    dups = df[dup_mask].copy()
    # Assign group numbers
    groups = dups.groupby(subset or list(df.columns)).ngroup()
    dups["duplicate_group"] = groups
    return dups.sort_values("duplicate_group")


def flag_missing(df: pd.DataFrame, threshold_pct: float = 50.0) -> dict[str, Any]:
    """Flag columns with high missing rates.

    Returns: {
        "total_rows": int,
        "columns_over_threshold": [{col, missing_pct, missing_count}, ...],
        "per_column": {col: {"missing_count": int, "missing_pct": float}, ...}
    }
    """
    total = len(df)
    per_column: dict[str, Any] = {}
    over: list[dict[str, Any]] = []
    for col in df.columns:
        missing = int(df[col].isna().sum())
        pct = round(missing / max(total, 1) * 100, 1)
        info = {"missing_count": missing, "missing_pct": pct}
        per_column[str(col)] = info
        if pct >= threshold_pct:
            over.append({"column": str(col), **info})
    return {"total_rows": total, "columns_over_threshold": over, "per_column": per_column}


def clean(
    path: str,
    operations: Optional[list[str]] = None,
    output_path: Optional[str] = None,
) -> dict[str, Any]:
    """Run common cleaning operations and save the result.

    Args:
        path: Input Excel file.
        operations: List of ops: 'trim', 'normalize_columns', 'normalize_dates',
                    'flag_duplicates', 'flag_missing'. Default: all except date parsing.
        output_path: Where to save. Auto-generated if None.

    Returns:
        {"cleaned_path": str, "changes": [...], "issues_remaining": [...]}
    """
    if operations is None:
        operations = ["trim", "normalize_columns", "flag_duplicates", "flag_missing"]

    all_sheets = read_excel(path, sheet=None)
    if not isinstance(all_sheets, dict):
        all_sheets = {"Sheet1": all_sheets}

    changes: list[str] = []
    issues: list[str] = []
    cleaned_sheets: dict[str, pd.DataFrame] = {}

    for name, df in all_sheets.items():
        df_clean = df.copy()

        if "trim" in operations:
            before_nulls = df_clean.isna().sum().sum()
            df_clean = trim_all_text(df_clean)
            changes.append(f"[{name}] trim: normalized text whitespace")

        if "normalize_columns" in operations:
            df_clean = normalize_column_names(df_clean)
            changes.append(f"[{name}] normalized column names")

        if "normalize_dates" in operations:
            date_cols = [
                c
                for c in df_clean.columns
                if any(
                    kw in str(c).lower()
                    for kw in ("date", "日期", "time", "时间", "day", "month", "year")
                )
            ]
            if date_cols:
                df_clean, date_report = normalize_dates(df_clean, date_cols)
                for col, info in date_report.items():
                    if "failed" in info and info["failed"] > 0:
                        issues.append(
                            f"[{name}] date column '{col}': {info['failed']} unparseable values"
                        )
                    else:
                        changes.append(
                            f"[{name}] date column '{col}': {info.get('parsed', '?')} parsed"
                        )

        if "flag_duplicates" in operations:
            dups = detect_duplicates(df_clean)
            if len(dups) > 0:
                issues.append(f"[{name}] {len(dups)} duplicate rows flagged")
            else:
                changes.append(f"[{name}] no duplicates found")

        if "flag_missing" in operations:
            missing_info = flag_missing(df_clean)
            over = missing_info["columns_over_threshold"]
            if over:
                for item in over:
                    issues.append(
                        f"[{name}] column '{item['column']}': "
                        f"{item['missing_pct']}% missing"
                    )

        cleaned_sheets[name] = df_clean

    if output_path is None:
        base = os.path.splitext(os.path.basename(path))[0]
        output_path = os.path.join(
            os.path.dirname(path) or ".",
            f"{base}_cleaned_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

    write_report(output_path, cleaned_sheets)

    return {
        "cleaned_path": os.path.abspath(output_path),
        "changes": changes,
        "issues_remaining": issues,
    }


# ── Section 3: Comparison ──────────────────────────────────────────────────


def compare_structure(*paths: str) -> dict[str, Any]:
    """Compare schema across 2+ files.

    Returns: {
        "files": [str, ...],
        "common_sheets": [str, ...],
        "unique_sheets": {path: [str, ...]},
        "per_sheet": {
            "<name>": {
                "common_columns": [str, ...],
                "unique_columns": {path: [str, ...]},
                "dtype_conflicts": {col: {path: dtype, ...}},
            }
        }
    }
    """
    all_info = {}
    for p in paths:
        all_info[p] = describe(p)

    file_names = [os.path.basename(p) for p in paths]
    sheet_sets = [set(all_info[p]["sheets"].keys()) for p in paths]
    common_sheets = list(set.intersection(*sheet_sets)) if sheet_sets else []
    unique_sheets = {
        file_names[i]: list(sheet_sets[i] - set(common_sheets))
        for i in range(len(paths))
    }

    per_sheet: dict[str, Any] = {}
    for sn in common_sheets:
        col_sets = {}
        dtype_maps = {}
        for i, p in enumerate(paths):
            info = all_info[p]["sheets"].get(sn, {})
            cols = set(info.get("column_names", []))
            col_sets[file_names[i]] = cols
            dtype_maps[file_names[i]] = info.get("dtypes", {})

        all_cols = list(set.union(*col_sets.values())) if col_sets else []
        common_cols = list(set.intersection(*col_sets.values())) if col_sets else []
        unique_cols = {
            fn: list(cols - set(common_cols)) for fn, cols in col_sets.items()
        }

        dtype_conflicts = {}
        for col in common_cols:
            dtypes = {fn: dtype_maps[fn].get(col, "N/A") for fn in file_names}
            if len(set(dtypes.values())) > 1:
                dtype_conflicts[col] = dtypes

        per_sheet[sn] = {
            "common_columns": common_cols,
            "unique_columns": unique_cols,
            "dtype_conflicts": dtype_conflicts,
        }

    return {
        "files": file_names,
        "common_sheets": common_sheets,
        "unique_sheets": unique_sheets,
        "per_sheet": per_sheet,
    }


def compare_values(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    key_columns: list[str],
    compare_columns: Optional[list[str]] = None,
    tolerance: float = 0.0,
) -> dict[str, Any]:
    """Row-level value comparison between two DataFrames.

    Returns: {
        "matched": DataFrame (rows in both, with _diff_<col> columns),
        "only_in_left": DataFrame,
        "only_in_right": DataFrame,
        "unchanged_count": int,
        "changed_count": int,
    }
    """
    # Ensure key columns exist
    for col in key_columns:
        if col not in df1.columns:
            raise ValueError(f"Key column '{col}' not found in left DataFrame")
        if col not in df2.columns:
            raise ValueError(f"Key column '{col}' not found in right DataFrame")

    if compare_columns is None:
        compare_columns = [c for c in df1.columns if c not in key_columns and c in df2.columns]

    left = df1.copy()
    right = df2.copy()

    # Tag source
    left["_source"] = "left"
    right["_source"] = "right"

    merged = pd.merge(
        left, right, on=key_columns, how="outer", suffixes=("_left", "_right"), indicator=True
    )

    only_left = merged[merged["_merge"] == "left_only"][
        key_columns + [c for c in left.columns if c != "_source"]
    ].drop(columns=["_source"], errors="ignore")

    only_right = merged[merged["_merge"] == "right_only"][
        key_columns + [c for c in right.columns if c != "_source"]
    ].drop(columns=["_source"], errors="ignore")

    both = merged[merged["_merge"] == "both"].copy()

    # Build diff columns
    changed_count = 0
    for col in compare_columns:
        col_l = f"{col}_left"
        col_r = f"{col}_right"
        if col_l not in both.columns or col_r not in both.columns:
            continue
        diff_col = f"_diff_{col}"
        # Try numeric comparison
        try:
            l_vals = pd.to_numeric(both[col_l], errors="coerce")
            r_vals = pd.to_numeric(both[col_r], errors="coerce")
            delta = (l_vals - r_vals).abs()
            both[diff_col] = both.apply(
                lambda row, c=col_l: f"{row.get(col_l, '?')} → {row.get(col_r, '?')}"
                if abs(
                    float(row.get(col_l, 0) or 0) - float(row.get(col_r, 0) or 0)
                ) > tolerance
                else "unchanged",
                axis=1,
            )
        except (ValueError, TypeError):
            # String comparison
            both[diff_col] = both.apply(
                lambda row, cl=col_l, cr=col_r: (
                    f"{row.get(cl, '?')} → {row.get(cr, '?')}"
                    if str(row.get(cl, "")) != str(row.get(cr, ""))
                    else "unchanged"
                ),
                axis=1,
            )
        changed_count += int((both[diff_col] != "unchanged").sum())

    changed_cols = [c for c in both.columns if c.startswith("_diff_")]
    result_cols = key_columns + compare_columns + changed_cols
    matched = both[result_cols]

    unchanged_count = len(matched) - changed_count // max(len(compare_columns), 1)

    return {
        "matched": matched,
        "only_in_left": only_left,
        "only_in_right": only_right,
        "unchanged_count": unchanged_count,
        "changed_rows": changed_count // max(len(compare_columns), 1),
    }


def compare(
    path_a: str,
    path_b: str,
    key_columns: Optional[list[str]] = None,
    sheet: Any = 0,
) -> dict[str, Any]:
    """High-level file comparison: structure diff + value diff.

    Returns a dict with 'structure' and 'values' keys.
    """
    structure = compare_structure(path_a, path_b)

    # If no key columns given, try to guess from common columns
    df_a = read_excel(path_a, sheet=sheet)
    df_b = read_excel(path_b, sheet=sheet)

    if key_columns is None:
        # Heuristic: first string column with unique values in both
        for col in df_a.columns:
            if col in df_b.columns:
                if df_a[col].dtype == object and df_a[col].nunique() == len(df_a):
                    if df_b[col].nunique() == len(df_b):
                        key_columns = [str(col)]
                        break
        if key_columns is None:
            key_columns = [str(df_a.columns[0])]

    values = compare_values(df_a, df_b, key_columns)

    return {"structure": structure, "values": values}


# ── Section 4: Statistical Profiling ───────────────────────────────────────


def profile_statistics(
    df: pd.DataFrame,
    metric_columns: Optional[list[str]] = None,
    group_by: Optional[list[str]] = None,
) -> pd.DataFrame:
    """Statistical profile of numeric columns.

    Returns a DataFrame with rows=metric, columns=count, mean, median, std,
    min, max, Q1, Q3, IQR, skew, kurtosis, missing, missing_pct.
    If group_by is set, returns MultiIndex grouped DataFrame.
    """
    if metric_columns is None:
        metric_columns = [
            c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])
        ]

    if group_by:
        results = []
        for name, group in df.groupby(group_by):
            stats = _profile_group(group, metric_columns)
            if isinstance(name, tuple):
                for i, gcol in enumerate(group_by):
                    stats[gcol] = name[i]
            else:
                stats[group_by[0]] = name
            results.append(stats)
        result = pd.DataFrame(results)
        cols = group_by + [c for c in result.columns if c not in group_by]
        return result[cols]

    return _profile_group(df, metric_columns)


def _profile_group(df: pd.DataFrame, metric_columns: list[str]) -> pd.DataFrame:
    rows = []
    for col in metric_columns:
        if col not in df.columns:
            continue
        series = df[col].dropna()
        vals = series.values
        if len(vals) == 0:
            rows.append({"metric": str(col), "count": 0, "missing": len(df), "missing_pct": 100.0})
            continue
        try:
            rows.append(
                {
                    "metric": str(col),
                    "count": len(vals),
                    "mean": round(float(np.mean(vals)), 4),
                    "median": round(float(np.median(vals)), 4),
                    "std": round(float(np.std(vals)), 4),
                    "min": float(np.min(vals)),
                    "max": float(np.max(vals)),
                    "Q1": round(float(np.percentile(vals, 25)), 4),
                    "Q3": round(float(np.percentile(vals, 75)), 4),
                    "IQR": round(float(np.percentile(vals, 75) - np.percentile(vals, 25)), 4),
                    "skew": round(float(pd.Series(vals).skew()), 4),
                    "missing": int(df[col].isna().sum()),
                    "missing_pct": round(df[col].isna().sum() / max(len(df), 1) * 100, 1),
                }
            )
        except Exception:
            rows.append({"metric": str(col), "count": len(vals), "error": "profiling failed"})
    return pd.DataFrame(rows)


def detect_outliers(
    df: pd.DataFrame,
    metric_columns: list[str],
    method: str = "iqr",
    threshold: float = 1.5,
) -> pd.DataFrame:
    """Flag outlier rows using IQR or Z-score method.

    Returns a DataFrame of flagged rows with columns:
        [original_index, column, value, lower_bound, upper_bound, deviation]
    """
    flagged: list[dict[str, Any]] = []
    for col in metric_columns:
        if col not in df.columns or not pd.api.types.is_numeric_dtype(df[col]):
            continue
        series = df[col].dropna()
        if len(series) < 4:
            continue

        if method == "iqr":
            q1 = float(np.percentile(series, 25))
            q3 = float(np.percentile(series, 75))
            iqr = q3 - q1
            lower = q1 - threshold * iqr
            upper = q3 + threshold * iqr
            outlier_mask = (df[col] < lower) | (df[col] > upper)
        else:  # zscore
            mean = float(np.mean(series))
            std = float(np.std(series))
            if std == 0:
                continue
            lower = mean - threshold * std
            upper = mean + threshold * std
            outlier_mask = (df[col] < lower) | (df[col] > upper)

        for idx in df.index[outlier_mask]:
            val = df.loc[idx, col]
            if pd.isna(val):
                continue
            flagged.append(
                {
                    "row_index": idx,
                    "column": str(col),
                    "value": float(val),  # type: ignore[arg-type]
                    "lower_bound": round(lower, 4),
                    "upper_bound": round(upper, 4),
                    "deviation": round(float(val) - float(np.mean(series)), 4),  # type: ignore[arg-type]
                }
            )

    return pd.DataFrame(flagged)


def detect_trends(
    df: pd.DataFrame,
    date_column: str,
    metric_columns: list[str],
) -> pd.DataFrame:
    """Linear trend detection per metric column against date.

    Returns DataFrame: metric, slope, intercept, r_squared, p_value, direction.
    """
    from scipy import stats as scipy_stats

    if date_column not in df.columns:
        raise ValueError(f"Date column '{date_column}' not found")

    dates = pd.to_datetime(df[date_column], errors="coerce")
    ordinal = dates.dropna().map(datetime.toordinal).astype(float)
    valid_idx = ordinal.index

    rows = []
    for col in metric_columns:
        if col not in df.columns or not pd.api.types.is_numeric_dtype(df[col]):
            continue
        y = df.loc[valid_idx, col].astype(float)
        mask = y.notna() & ordinal.notna()
        x = ordinal[mask]
        y = y[mask]
        if len(y) < 3:
            rows.append({"metric": col, "error": "fewer than 3 data points"})
            continue
        try:
            slope, intercept, r_value, p_value, _ = scipy_stats.linregress(x, y)
            rows.append(
                {
                    "metric": str(col),
                    "slope": round(float(slope), 6),
                    "intercept": round(float(intercept), 4),
                    "r_squared": round(float(r_value**2), 4),
                    "p_value": round(float(p_value), 6),
                    "direction": "up" if slope > 0 else ("down" if slope < 0 else "flat"),
                    "significant": bool(p_value < 0.05),
                }
            )
        except Exception:
            rows.append({"metric": col, "error": "regression failed"})

    return pd.DataFrame(rows)


def profile(
    path: str,
    columns: Optional[list[str]] = None,
    sheet: Any = 0,
) -> dict[str, Any]:
    """High-level profiling entry point for CLI.

    Returns describe() output + stats + outliers in one package.
    """
    desc = describe(path)
    df = read_excel(path, sheet=sheet)
    if not isinstance(df, pd.DataFrame):
        df = list(df.values())[0] if df else pd.DataFrame()

    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if columns:
        numeric_cols = [c for c in columns if c in numeric_cols]

    stats = profile_statistics(df, numeric_cols)
    outliers = detect_outliers(df, numeric_cols) if numeric_cols else pd.DataFrame()

    stats_dict = stats.to_dict(orient="records") if not stats.empty else []
    outlier_dict = outliers.to_dict(orient="records") if not outliers.empty else []

    return {
        "describe": desc,
        "statistics": stats_dict,
        "outlier_count": len(outlier_dict),
        "outliers": outlier_dict[:50],  # cap at 50 for context window
    }


# ── Section 5: Writing ──────────────────────────────────────────────────────


def write_report(
    output_path: str,
    sheets: dict[str, pd.DataFrame],
    autoformat: bool = True,
) -> str:
    """Write a multi-sheet Excel report.

    Args:
        output_path: Absolute or relative output path.
        sheets: dict of sheet_name → DataFrame.
        autoformat: Apply column-width, freeze-panes, and auto-filter.

    Returns:
        Absolute path to the written file.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            sheet_name = name[:31]  # Excel 31-char sheet name limit
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    if autoformat:
        try:
            _autoformat(output_path)
        except Exception:
            pass  # autoformat is cosmetic, never fail the write

    return os.path.abspath(output_path)


def _autoformat(path: str) -> None:
    """Apply sensible defaults: auto-width, freeze header, auto-filter."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        # Auto-width
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        # Freeze header row
        ws.freeze_panes = "A2"
        # Auto-filter
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
    wb.save(path)


# ── Section 6: CLI ──────────────────────────────────────────────────────────


def _cli_describe(args: argparse.Namespace) -> None:
    result = describe(args.file)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


def _cli_clean(args: argparse.Namespace) -> None:
    ops = args.operations.split(",") if args.operations else None
    result = clean(args.file, operations=ops, output_path=args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


def _cli_compare(args: argparse.Namespace) -> None:
    key_cols = args.key.split(",") if args.key else None
    result = compare(args.file_a, args.file_b, key_columns=key_cols)
    # DataFrames inside; serialize
    serializable = {
        "structure": result["structure"],
        "values": {
            "only_in_left_count": len(result["values"]["only_in_left"]),
            "only_in_right_count": len(result["values"]["only_in_right"]),
            "matched_count": len(result["values"]["matched"]),
            "unchanged_count": result["values"]["unchanged_count"],
            "changed_rows": result["values"]["changed_rows"],
            "only_in_left_preview": result["values"]["only_in_left"]
            .head(10)
            .to_dict(orient="records"),
            "only_in_right_preview": result["values"]["only_in_right"]
            .head(10)
            .to_dict(orient="records"),
            "changed_preview": result["values"]["matched"]
            .head(20)
            .to_dict(orient="records"),
        },
    }
    print(json.dumps(serializable, ensure_ascii=False, indent=2, default=str))


def _cli_profile(args: argparse.Namespace) -> None:
    cols = args.columns.split(",") if args.columns else None
    result = profile(args.file, columns=cols)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="excel_core.py — Claude Code + Excel utility library"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_desc = sub.add_parser("describe", help="Structural summary of a workbook")
    p_desc.add_argument("file", help="Path to Excel/CSV file")

    p_clean = sub.add_parser("clean", help="Run cleaning operations")
    p_clean.add_argument("file", help="Path to Excel file")
    p_clean.add_argument("--output", "-o", default=None, help="Output path")
    p_clean.add_argument(
        "--operations",
        default="trim,normalize_columns,flag_duplicates,flag_missing",
        help="Comma-separated list of operations",
    )

    p_comp = sub.add_parser("compare", help="Compare two Excel files")
    p_comp.add_argument("file_a", help="First file")
    p_comp.add_argument("file_b", help="Second file")
    p_comp.add_argument("--key", "-k", default=None, help="Key column(s), comma-separated")

    p_prof = sub.add_parser("profile", help="Statistical profile + outliers")
    p_prof.add_argument("file", help="Path to Excel file")
    p_prof.add_argument("--columns", "-c", default=None, help="Metric columns, comma-separated")

    args = parser.parse_args()

    dispatch = {
        "describe": _cli_describe,
        "clean": _cli_clean,
        "compare": _cli_compare,
        "profile": _cli_profile,
    }
    dispatch[args.command](args)


if __name__ == "__main__":
    main()
